import pandas as pd
import openpyxl
import csv


# Example usage
CSV_FILE_PATH = 'result_csv_file.csv'
EXCEL_FILE_PATH = 'data.xlsx'
SHEET_NAME = 'Basic Worksheet'


def write_csv(file_path, data, headers="None"):
    print(data)
    with open(file_path, 'a', newline='') as csvfile:
        writer = csv.writer(csvfile)

        # Write headers if provided
        if headers:
            writer.writerow(headers)

        # Write data rows
        for i in data:
            writer.writerows(i)


def validate_data(data):
    pattern_first_name = r'[A-Za-z]+'
    pattern_last_name = r'[A-Za-z]+'
    pattern_ssn = r'\d{3}-\d{2}-\d{4}'
    pattern_address = r'.+'
    pattern_company = r'.+'
    pattern_department = r'.+'
    pattern_position = r'.+'
    pattern_zip = r'\d{5}(-\d{4})?'
    pattern_mobile_number = r'\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}'


def read_xlsx(filename: str):
    df = pd.read_excel(filename)
    data_as_list_of_dicts = df.to_dict('records')
    return data_as_list_of_dicts
    # print(data_as_list_of_dicts)
    # for _dict in data_as_list_of_dicts:
    #     yield _dict


def read_excel_line_by_line(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    # Get the maximum number of rows and columns in the sheet
    max_row = sheet.max_row
    max_column = sheet.max_column

    # Iterate through each row in the sheet
    for row_num in range(1, max_row + 1):
        row_data = []

        # Iterate through each column in the row
        for col_num in range(1, max_column + 1):
            cell_value = sheet.cell(row=row_num, column=col_num).value
            row_data.append(cell_value)

        yield row_data


def main():
    # for row in read_excel_line_by_line(EXCEL_FILE_PATH, SHEET_NAME):
    #     write_csv(CSV_FILE_PATH, row)
    read_xlsx(EXCEL_FILE_PATH)
    df = pd.DataFrame(read_xlsx(EXCEL_FILE_PATH))
    df.to_csv("result_csv_file.csv")


if __name__ == '__main__':
    main()

